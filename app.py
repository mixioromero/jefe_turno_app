from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from datetime import datetime, date, time
from pathlib import Path
import json
import os

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / 'data.xlsx'
STATIC_DIR = BASE_DIR / 'static'

app = Flask(__name__)
app.secret_key = os.environ.get('APP_SECRET_KEY', 'jefe-turno-puerto')

SHEETS_CONFIG = {
    'Bloqueos Activos': {'table': 'Bloqueos_Activos', 'title': 'Bloqueos Activos', 'icon': '🔒'},
    'ART': {'table': 'ART', 'title': 'ART', 'icon': '🦺'},
    'Checklist Moviles': {'table': 'Checklist_Moviles', 'title': 'Checklist Móviles', 'icon': '🚜'},
    'Checklist Fijos': {'table': 'Checklist_Fijos', 'title': 'Checklist Fijos', 'icon': '🏗️'},
    'PTS': {'table': 'PTS', 'title': 'PTS', 'icon': '📘'},
    'Anomalías': {'table': 'Anomalías', 'title': 'Anomalías', 'icon': '⚠️'},
    'Avisos SAP': {'table': 'Avisos_SAP', 'title': 'Avisos SAP', 'icon': '🛠️'},
    'Bitácora Turno': {'table': 'Bitácora_Turno', 'title': 'Bitácora Turno', 'icon': '📝'},
    'Entrega Turno': {'table': 'Entrega_Turno', 'title': 'Entrega de Turno', 'icon': '🤝'},
    'Asignacion VHF': {'table': 'Asignacion_VHF', 'title': 'Asignación VHF', 'icon': '📻'},
}

MODULE_GROUPS = {
    'Operación y control': ['Bloqueos Activos', 'Bitácora Turno', 'Entrega Turno', 'Asignacion VHF'],
    'Seguridad y permisos': ['ART', 'PTS', 'Anomalías'],
    'Inspecciones y mantenimiento': ['Checklist Moviles', 'Checklist Fijos', 'Avisos SAP'],
}

MASTER_TO_OPTIONS = {
    'Estado': 'Estados',
    'Estado final': 'Estados',
    'Prioridad': 'Prioridades',
    'Tipo': 'Tipos_Anomalia',
    'Turno': 'Turnos',
    'Turno saliente': 'Turnos',
    'Turno entrante': 'Turnos',
    'Combustible': 'Estados_OK',
    'Aceite motor': 'Estados_OK',
    'Aceite hidráulico': 'Estados_OK',
    'Refrigerante': 'Estados_OK',
    'Luces': 'Estados_OK',
    'Alarma': 'Estados_OK',
    'Neumáticos': 'Estados_OK',
    'Fugas': 'Estados_OK',
    'Cabina': 'Estados_OK',
    'Frenos': 'Estados_OK',
    'Extintor': 'Estados_OK',
    'Estructura': 'Estados_OK',
    'Motor': 'Estados_OK',
    'Guardas': 'Estados_OK',
    'Sensores': 'Estados_OK',
    'Polines': 'Estados_OK',
    'Correas': 'Estados_OK',
    'Chutes': 'Estados_OK',
    'Tolvas': 'Estados_OK',
    'Alimentadores': 'Estados_OK',
    'Captadores': 'Estados_OK',
    'Nebulizador': 'Estados_OK',
    'Impacto': 'Impactos',
}

DATE_FIELDS = {'Fecha', 'Cierre', 'Próx revisión'}
TIME_FIELDS = {'Hora', 'Hora liberación', 'Hora Entrega', 'Hora Devolución'}
NUMERIC_FIELDS = {'Personal', 'Canal', 'Horómetro', 'Equipos OK', 'Equipos falla', 'Bloqueos', 'Avisos SAP'}
TEXTAREA_FIELDS = {'Motivo', 'Obs', 'Observación', 'Descripción', 'Acción', 'Comentarios', 'Pendientes', 'Recomendaciones', 'Observaciones', 'Riesgos', 'Consecuencias', 'Medidas'}
AUTONUMERIC_FIELDS = {'ID'}
DETAIL_EXCLUDED = {'row_num', 'search_blob'}
PREFERRED_CLOSE_VALUES = ['Cerrada', 'Cerrado', 'Finalizada', 'Finalizado', 'Resuelto', 'Resuelta', 'Completado', 'Completada', 'OK', 'Liberado']
OPEN_TOKENS = {'activo', 'abierto', 'pendiente', 'en proceso', 'bloqueado', 'observado'}
CLOSED_TOKENS = {'cerrado', 'cerrada', 'resuelto', 'resuelta', 'finalizado', 'finalizada', 'ok', 'completado', 'completada', 'liberado'}
SUMMARY_PRIORITY = [
    'Equipo', 'Código', 'Aviso SAP', 'N° SAP', 'Área', 'Evento', 'Actividad', 'Descripción', 'Responsable', 'Prioridad',
    'Estado', 'Estado final', 'Turno', 'Hora', 'Fecha'
]
TITLE_FIELDS = ['Equipo', 'Evento', 'Actividad', 'Procedimiento', 'Código VHF', 'Aviso SAP', 'Descripción', 'Área']
SUBTITLE_FIELDS = ['Área', 'Turno', 'Responsable', 'Prioridad', 'Estado', 'Estado final', 'Fecha', 'Hora']


def load_wb():
    return load_workbook(EXCEL_PATH)



def get_headers(ws):
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]



def get_master_lists(wb):
    ws = wb['MAESTRO_LISTAS']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    lists = {}
    for idx, header in enumerate(headers, start=1):
        values = []
        for r in range(2, ws.max_row + 1):
            value = ws.cell(r, idx).value
            if value not in (None, ''):
                values.append(str(value))
        lists[header] = values
    return lists



def dashboard_counts(wb):
    counts = {}
    checks = [
        ('Bloqueos Activos', 'Estado', lambda v: str(v).strip().lower() == 'activo'),
        ('Anomalías', 'Estado', lambda v: str(v).strip().lower() != 'cerrada'),
        ('Avisos SAP', 'Estado', lambda v: str(v).strip().lower() == 'abierto'),
        ('Bitácora Turno', 'Hora', lambda v: v not in (None, '')),
    ]
    for sheet_name, field, predicate in checks:
        ws = wb[sheet_name]
        headers = get_headers(ws)
        idx = headers.index(field) + 1
        total = 0
        for r in range(2, ws.max_row + 1):
            value = ws.cell(r, idx).value
            if predicate(value):
                total += 1
        counts[sheet_name] = total
    return counts



def infer_input_type(header):
    if header in AUTONUMERIC_FIELDS:
        return 'autonumber'
    if header in DATE_FIELDS:
        return 'date'
    if header in TIME_FIELDS:
        return 'time'
    if header in NUMERIC_FIELDS:
        return 'number'
    if header in TEXTAREA_FIELDS:
        return 'textarea'
    return 'text'



def next_id_value(ws, header):
    headers = get_headers(ws)
    idx = headers.index(header) + 1
    max_id = 0
    for r in range(2, ws.max_row + 1):
        value = ws.cell(r, idx).value
        digits = ''.join(ch for ch in str(value or '') if ch.isdigit())
        if digits:
            max_id = max(max_id, int(digits))
    if header == 'ID':
        return f'A-{max_id + 1:03d}'
    return max_id + 1



def parse_value(header, value, ws=None, preserve_existing_id=False):
    if header in AUTONUMERIC_FIELDS and ws is not None and not preserve_existing_id:
        return next_id_value(ws, header)
    if value is None:
        return None
    value = value.strip() if isinstance(value, str) else value
    if value == '':
        return None
    if header in DATE_FIELDS:
        parsed = try_parse_date(value)
        return parsed if parsed is not None else value
    if header in TIME_FIELDS:
        parsed = try_parse_time(value)
        return parsed if parsed is not None else value
    if header in NUMERIC_FIELDS:
        try:
            return int(value)
        except Exception:
            try:
                return float(value)
            except Exception:
                return value
    return value



def append_row_and_expand_table(ws, table_name, row_data):
    next_row = ws.max_row + 1
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(next_row, col_idx, value)

    table = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    if next_row > max_row:
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{next_row}"



def today_iso():
    return datetime.now().strftime('%Y-%m-%d')



def now_hhmm():
    return datetime.now().strftime('%H:%M')



def get_default_value(header, ws=None):
    if header in AUTONUMERIC_FIELDS and ws is not None:
        return str(next_id_value(ws, header))
    if header in DATE_FIELDS:
        return today_iso()
    if header in TIME_FIELDS:
        return now_hhmm()
    return ''



def format_for_display(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d-%m-%Y %H:%M')
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime('%d-%m-%Y')
    if isinstance(value, time):
        return value.strftime('%H:%M')
    return str(value)



def try_parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                continue
    return None



def try_parse_time(value):
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                return datetime.strptime(raw, fmt).time().replace(second=0, microsecond=0)
            except Exception:
                continue
    return None



def format_for_input(header, value):
    if value in (None, ''):
        return ''
    if header in DATE_FIELDS:
        parsed = try_parse_date(value)
        return parsed.strftime('%Y-%m-%d') if parsed else str(value)
    if header in TIME_FIELDS:
        parsed = try_parse_time(value)
        return parsed.strftime('%H:%M') if parsed else str(value)
    return str(value)



def get_status_field(headers):
    for candidate in ['Estado', 'Estado final']:
        if candidate in headers:
            return candidate
    return None



def normalize_status(value):
    return str(value or '').strip().lower()



def is_closed_status(value):
    normalized = normalize_status(value)
    return normalized in CLOSED_TOKENS



def is_open_status(value):
    normalized = normalize_status(value)
    if not normalized:
        return False
    return normalized in OPEN_TOKENS



def status_badge_class(value):
    normalized = normalize_status(value)
    if normalized in CLOSED_TOKENS:
        return 'is-closed'
    if normalized in OPEN_TOKENS:
        return 'is-open'
    return 'is-neutral'



def available_existing_values(ws, field):
    headers = get_headers(ws)
    if field not in headers:
        return []
    idx = headers.index(field) + 1
    seen = []
    for row_num in range(2, ws.max_row + 1):
        value = ws.cell(row_num, idx).value
        text = str(value).strip() if value not in (None, '') else ''
        if text and text not in seen:
            seen.append(text)
    return seen



def get_close_value(sheet_name, headers, masters, ws):
    state_field = get_status_field(headers)
    if not state_field:
        return None, None
    options = masters.get(MASTER_TO_OPTIONS.get(state_field, ''), []) + available_existing_values(ws, state_field)
    dedup = []
    for opt in options:
        if opt not in dedup:
            dedup.append(opt)
    for preferred in PREFERRED_CLOSE_VALUES:
        for option in dedup:
            if option.strip().lower() == preferred.lower():
                return state_field, option
    if sheet_name in {'Anomalías', 'Avisos SAP'}:
        return state_field, 'Cerrada' if sheet_name == 'Anomalías' else 'Cerrado'
    return None, None



def fields_metadata(headers, masters, ws, mode='create'):
    fields = []
    for header in headers:
        input_type = infer_input_type(header)
        fields.append({
            'name': header,
            'label': header,
            'type': input_type,
            'options': masters.get(MASTER_TO_OPTIONS.get(header, ''), []),
            'default': get_default_value(header, ws) if mode == 'create' else '',
            'readonly': input_type == 'autonumber' and mode == 'create',
            'readonly_edit': header in AUTONUMERIC_FIELDS,
        })
    return fields



def build_summary_items(display_row):
    chosen = []
    used = set()
    for key in SUMMARY_PRIORITY:
        value = display_row.get(key)
        if value not in (None, '') and key not in used:
            chosen.append((key, value))
            used.add(key)
        if len(chosen) >= 4:
            return chosen
    for key, value in display_row.items():
        if value not in (None, '') and key not in used and key not in {'Obs', 'Observación', 'Descripción', 'Comentarios', 'Acción', 'Riesgos', 'Medidas'}:
            chosen.append((key, value))
            if len(chosen) >= 4:
                break
    return chosen



def build_row_title(display_row, row_num):
    for key in TITLE_FIELDS:
        value = display_row.get(key)
        if value not in (None, ''):
            return value
    return f'Registro #{row_num}'



def build_subtitle(display_row):
    parts = []
    for key in SUBTITLE_FIELDS:
        value = display_row.get(key)
        if value not in (None, ''):
            parts.append(f'{key}: {value}')
        if len(parts) >= 3:
            break
    return ' · '.join(parts)



def build_detail_items(display_row):
    items = []
    for key, value in display_row.items():
        if value not in (None, ''):
            items.append((key, value))
    return items



def build_search_blob(display_row):
    return ' '.join(str(v).lower() for v in display_row.values() if v not in (None, ''))



def sheet_rows(ws, headers, close_action=None):
    rows = []
    state_field = get_status_field(headers)
    for row_num in range(2, ws.max_row + 1):
        values = [ws.cell(row_num, c).value for c in range(1, len(headers) + 1)]
        if not any(v not in (None, '') for v in values):
            continue
        display_row = {header: format_for_display(value) for header, value in zip(headers, values)}
        raw_row = {header: format_for_input(header, value) for header, value in zip(headers, values)}
        state_value = display_row.get(state_field, '') if state_field else ''
        rows.append({
            'row_num': row_num,
            'title': build_row_title(display_row, row_num),
            'subtitle': build_subtitle(display_row),
            'summary_items': build_summary_items(display_row),
            'detail_items': build_detail_items(display_row),
            'display': display_row,
            'raw': raw_row,
            'state_field': state_field,
            'state_value': state_value,
            'state_badge_class': status_badge_class(state_value),
            'is_closed': is_closed_status(state_value),
            'can_close': bool(close_action and state_field and not is_closed_status(state_value)),
            'search_blob': build_search_blob(display_row),
        })
    return rows



def group_records(rows):
    open_rows = [row for row in rows if not row['is_closed']]
    closed_rows = [row for row in rows if row['is_closed']]
    if open_rows and closed_rows:
        return [
            {'key': 'abiertos', 'title': 'Abiertos / pendientes', 'rows': open_rows},
            {'key': 'cerrados', 'title': 'Cerrados / resueltos', 'rows': closed_rows},
        ]
    return [{'key': 'todos', 'title': 'Registros', 'rows': rows}]



def module_stats(rows):
    total = len(rows)
    open_count = len([row for row in rows if not row['is_closed']])
    closed_count = len([row for row in rows if row['is_closed']])
    return {
        'total': total,
        'open': open_count,
        'closed': closed_count,
        'has_state': any(row['state_field'] for row in rows),
    }



def bloqueo_estado_ui(value):
    estado = str(value or '').strip().lower()
    return 'Bloqueado' if estado in {'activo', 'bloqueado', 'sí', 'si'} else 'No bloqueado'



def bloqueo_estado_excel(value):
    return 'Activo' if value == 'Bloqueado' else 'Liberado'



def bloqueo_badge_class(value):
    return 'is-blocked' if value == 'Bloqueado' else 'is-clear'



def bloqueo_bucket(area):
    area_text = str(area or '').strip().lower()
    if 'desembarque' in area_text:
        return 'Desembarque'
    if 'embarque' in area_text:
        return 'Embarque'
    return 'Otros'



def bloqueos_groups(rows):
    groups = {'Embarque': [], 'Desembarque': [], 'Otros': []}
    for row in rows:
        groups.setdefault(row['bucket'], []).append(row)
    return groups



def bloqueos_rows(ws):
    headers = get_headers(ws)
    idx = {header: headers.index(header) + 1 for header in headers}
    rows = []
    for row_num in range(2, ws.max_row + 1):
        equipo = ws.cell(row_num, idx['Equipo']).value
        area = ws.cell(row_num, idx['Área']).value
        if not equipo and not area:
            continue
        estado_ui = bloqueo_estado_ui(ws.cell(row_num, idx['Estado']).value)
        tipo_bloqueo = ws.cell(row_num, idx['Tipo bloqueo']).value or ''
        personal = 'Eléctrico' if str(tipo_bloqueo).strip().lower() in {'', 'mecánico', 'mecanico'} else str(tipo_bloqueo)
        motivo = ws.cell(row_num, idx['Motivo']).value or ''
        obs = ws.cell(row_num, idx['Obs']).value or ''
        comentario = motivo or obs or 'Sin comentario registrado'
        rows.append({
            'row_num': row_num,
            'bucket': bloqueo_bucket(area),
            'area': area or 'Sin área',
            'equipo': equipo or 'Sin equipo',
            'tipo_bloqueo': tipo_bloqueo,
            'personal': personal,
            'responsable': ws.cell(row_num, idx['Responsable']).value or '',
            'estado_ui': estado_ui,
            'estado_badge': bloqueo_badge_class(estado_ui),
            'motivo': motivo,
            'obs': obs,
            'comentario': comentario,
        })
    return rows



def build_dashboard(wb):
    counts = dashboard_counts(wb)
    modules_map = {}
    for sheet_name, cfg in SHEETS_CONFIG.items():
        href = url_for('bloqueos_activos') if sheet_name == 'Bloqueos Activos' else url_for('form_sheet', sheet_name=sheet_name)
        modules_map[sheet_name] = {
            'sheet_name': sheet_name,
            'title': cfg['title'],
            'icon': cfg['icon'],
            'count': counts.get(sheet_name, ''),
            'href': href,
            'accent': 'danger' if sheet_name in {'Bloqueos Activos', 'Anomalías'} else 'brand',
        }

    sections = []
    for title, items in MODULE_GROUPS.items():
        sections.append({
            'title': title,
            'items': [modules_map[name] for name in items if name in modules_map],
        })

    flat_modules = [modules_map[name] for name in SHEETS_CONFIG.keys()]
    kpis = {
        'bloqueos': counts.get('Bloqueos Activos', 0),
        'anomalias': counts.get('Anomalías', 0),
        'avisos': counts.get('Avisos SAP', 0),
        'bitacora': counts.get('Bitácora Turno', 0),
    }
    return sections, flat_modules, kpis


@app.route('/')
def index():
    wb = load_wb()
    sections, modules, kpis = build_dashboard(wb)
    return render_template('index.html', sections=sections, modules=modules, kpis=kpis)


@app.route('/bloqueos-activos', methods=['GET', 'POST'])
def bloqueos_activos():
    wb = load_wb()
    ws = wb['Bloqueos Activos']
    headers = get_headers(ws)
    idx_estado = headers.index('Estado') + 1
    idx_motivo = headers.index('Motivo') + 1
    idx_hora_liberacion = headers.index('Hora liberación') + 1
    idx_tipo_bloqueo = headers.index('Tipo bloqueo') + 1

    if request.method == 'POST':
        row_num = int(request.form.get('row_num'))
        estado_ui = request.form.get('estado_ui', 'No bloqueado')
        motivo = (request.form.get('motivo') or '').strip()

        ws.cell(row_num, idx_estado, bloqueo_estado_excel(estado_ui))
        ws.cell(row_num, idx_motivo, motivo if motivo else None)
        ws.cell(row_num, idx_tipo_bloqueo, 'Eléctrico')
        if estado_ui == 'No bloqueado':
            ws.cell(row_num, idx_hora_liberacion, datetime.now().time().replace(second=0, microsecond=0))
        else:
            ws.cell(row_num, idx_hora_liberacion, None)

        wb.save(EXCEL_PATH)
        flash('Bloqueo actualizado correctamente.', 'success')
        return redirect(url_for('bloqueos_activos', selected=row_num))

    rows = bloqueos_rows(ws)
    selected = request.args.get('selected', type=int)

    return render_template(
        'bloqueos.html',
        title='Bloqueos Activos',
        icon='🔒',
        rows=rows,
        groups=bloqueos_groups(rows),
        selected_row=selected,
    )


@app.route('/form/<path:sheet_name>', methods=['GET', 'POST'])
def form_sheet(sheet_name):
    wb = load_wb()
    if sheet_name not in SHEETS_CONFIG:
        flash('Módulo no encontrado.', 'error')
        return redirect(url_for('index'))

    ws = wb[sheet_name]
    headers = get_headers(ws)
    masters = get_master_lists(wb)

    if request.method == 'POST':
        row_data = []
        for header in headers:
            raw = request.form.get(header, '')
            row_data.append(parse_value(header, raw, ws))
        append_row_and_expand_table(ws, SHEETS_CONFIG[sheet_name]['table'], row_data)
        wb.save(EXCEL_PATH)
        flash(f'Registro guardado en {SHEETS_CONFIG[sheet_name]["title"]}.', 'success')
        return redirect(url_for('form_sheet', sheet_name=sheet_name))

    rows = sheet_rows(ws, headers, get_close_value(sheet_name, headers, masters, ws))
    state_field, close_value = get_close_value(sheet_name, headers, masters, ws)
    return render_template(
        'form.html',
        sheet_name=sheet_name,
        title=SHEETS_CONFIG[sheet_name]['title'],
        icon=SHEETS_CONFIG[sheet_name]['icon'],
        fields=fields_metadata(headers, masters, ws, mode='create'),
        edit_fields=fields_metadata(headers, masters, ws, mode='edit'),
        rows=rows,
        grouped_rows=group_records(rows),
        stats=module_stats(rows),
        close_action={'field': state_field, 'value': close_value} if state_field and close_value else None,
    )


@app.post('/form/<path:sheet_name>/edit/<int:row_num>')
def edit_sheet_row(sheet_name, row_num):
    wb = load_wb()
    if sheet_name not in SHEETS_CONFIG:
        flash('Módulo no encontrado.', 'error')
        return redirect(url_for('index'))

    ws = wb[sheet_name]
    headers = get_headers(ws)
    if row_num < 2 or row_num > ws.max_row:
        flash('Registro no encontrado.', 'error')
        return redirect(url_for('form_sheet', sheet_name=sheet_name))

    for col_idx, header in enumerate(headers, start=1):
        preserve_id = header in AUTONUMERIC_FIELDS
        raw = request.form.get(header, '')
        value = parse_value(header, raw, ws, preserve_existing_id=preserve_id)
        if preserve_id and raw not in (None, ''):
            value = raw
        ws.cell(row_num, col_idx, value)

    wb.save(EXCEL_PATH)
    flash('Registro actualizado correctamente.', 'success')
    return redirect(url_for('form_sheet', sheet_name=sheet_name))


@app.post('/form/<path:sheet_name>/close/<int:row_num>')
def close_sheet_row(sheet_name, row_num):
    wb = load_wb()
    if sheet_name not in SHEETS_CONFIG:
        flash('Módulo no encontrado.', 'error')
        return redirect(url_for('index'))

    ws = wb[sheet_name]
    headers = get_headers(ws)
    masters = get_master_lists(wb)
    state_field, close_value = get_close_value(sheet_name, headers, masters, ws)
    if not state_field or not close_value:
        flash('Este módulo no tiene cierre rápido configurado.', 'error')
        return redirect(url_for('form_sheet', sheet_name=sheet_name))

    state_idx = headers.index(state_field) + 1
    ws.cell(row_num, state_idx, close_value)

    if 'Cierre' in headers:
        cierre_idx = headers.index('Cierre') + 1
        existing = ws.cell(row_num, cierre_idx).value
        if existing in (None, ''):
            ws.cell(row_num, cierre_idx, datetime.now().date())

    wb.save(EXCEL_PATH)
    flash('Registro cerrado correctamente.', 'success')
    return redirect(url_for('form_sheet', sheet_name=sheet_name))


@app.route('/manifest.webmanifest')
def manifest():
    return send_from_directory(STATIC_DIR, 'manifest.webmanifest', mimetype='application/manifest+json')


@app.route('/service-worker.js')
def service_worker():
    return send_from_directory(STATIC_DIR, 'service-worker.js', mimetype='application/javascript')


@app.route('/health')
def health():
    return {'status': 'ok'}
